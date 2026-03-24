import argparse, re, sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except Exception:
    sys.stderr.write("Missing dependency: openpyxl\n")
    sys.stderr.write("Install: python3 -m pip install openpyxl\n")
    raise


def _escape_md_cell(v):
    s = "" if v is None else str(v)
    s = s.replace("\\", "\\\\")
    s = s.replace("|", "\\|")
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")
    return s


def _is_blank_cell_value(v):
    if v is None:
        return True
    if isinstance(v, str):
        return v.strip() == ""
    return False


def _format_number(v):
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        s = str(v)
        if "e" in s.lower():
            s = format(v, "f").rstrip("0").rstrip(".")
        else:
            if "." in s:
                s = s.rstrip("0").rstrip(".")
        return s
    return str(v)


def _format_cell(cell):
    v = cell.value
    if v is None:
        return ""
    if isinstance(v, datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0 and v.microsecond == 0:
            return v.strftime("%Y-%m-%d")
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float, bool)):
        return _format_number(v)
    return str(v)


def _build_merged_value_map(ws):
    m = {}
    for r in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = r.min_row, r.min_col, r.max_row, r.max_col
        v = ws.cell(row=min_row, column=min_col).value
        if v is None or str(v).strip() == "":
            continue
        for rr in range(min_row, max_row + 1):
            for cc in range(min_col, max_col + 1):
                m[(rr, cc)] = v
    return m


def _find_used_bounds(ws, merged_map):
    min_row = None
    min_col = None
    max_row = None
    max_col = None

    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if v is None or str(v).strip() == "":
                continue
            r, c = cell.row, cell.column
            min_row = r if min_row is None else min(min_row, r)
            min_col = c if min_col is None else min(min_col, c)
            max_row = r if max_row is None else max(max_row, r)
            max_col = c if max_col is None else max(max_col, c)

    for (r, c), v in merged_map.items():
        if v is None or str(v).strip() == "":
            continue
        min_row = r if min_row is None else min(min_row, r)
        min_col = c if min_col is None else min(min_col, c)
        max_row = r if max_row is None else max(max_row, r)
        max_col = c if max_col is None else max(max_col, c)

    if min_row is None:
        return None
    return (min_row, min_col, max_row, max_col)


def _trim_matrix_outer_empty(matrix):
    if not matrix:
        return matrix
    row_has = [any(str(c).strip() != "" for c in row) for row in matrix]
    if not any(row_has):
        return []

    top = next(i for i, ok in enumerate(row_has) if ok)
    bottom = len(row_has) - 1 - next(i for i, ok in enumerate(reversed(row_has)) if ok)
    matrix = matrix[top:bottom + 1]

    col_count = max(len(r) for r in matrix)
    norm = [r + [""] * (col_count - len(r)) for r in matrix]
    col_has = [any(str(norm[r][c]).strip() != "" for r in range(len(norm))) for c in range(col_count)]
    left = next(i for i, ok in enumerate(col_has) if ok)
    right = len(col_has) - 1 - next(i for i, ok in enumerate(reversed(col_has)) if ok)

    return [r[left:right + 1] for r in norm]


def _sheet_to_matrix(ws, trim_outer=True, max_cells=20000, max_rows=None, max_cols=None):
    merged_map = _build_merged_value_map(ws)
    bounds = _find_used_bounds(ws, merged_map)
    if bounds is None:
        return []
    min_row, min_col, max_row, max_col = bounds
    row_count = max_row - min_row + 1
    col_count = max_col - min_col + 1
    if row_count * col_count > max_cells:
        raise ValueError(f"sheet_too_large: {ws.title} ({row_count}x{col_count}={row_count*col_count} cells)")

    matrix = []
    for r in range(min_row, max_row + 1):
        row = []
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            if _is_blank_cell_value(cell.value) and (r, c) in merged_map:
                row.append("" if merged_map[(r, c)] is None else str(merged_map[(r, c)]))
            else:
                row.append(_format_cell(cell))
        matrix.append(row)

    if trim_outer:
        matrix = _trim_matrix_outer_empty(matrix)

    if max_rows is not None and max_rows >= 0:
        matrix = matrix[:max_rows]
    if max_cols is not None and max_cols >= 0:
        matrix = [r[:max_cols] for r in matrix]
    return matrix


def _is_blank_row(row):
    return all(str(c).strip() == "" for c in row)


def _split_tables_by_blank_rows(matrix, blank_rows_gap=1, min_rows=2):
    if not matrix:
        return []
    tables = []
    cur = []
    blank_run = 0
    for row in matrix:
        if _is_blank_row(row):
            blank_run += 1
            if cur and blank_run >= blank_rows_gap:
                tables.append(cur)
                cur = []
            continue
        blank_run = 0
        cur.append(row)
    if cur:
        tables.append(cur)
    tables = [_trim_matrix_outer_empty(t) for t in tables]
    tables = [t for t in tables if len(t) >= min_rows and any(not _is_blank_row(r) for r in t)]
    return tables


def _matrix_to_markdown_table(matrix, header_row=None, no_header=False):
    if not matrix:
        return ""
    col_count = max(len(r) for r in matrix)
    rows = [r + [""] * (col_count - len(r)) for r in matrix]

    if no_header:
        header = [f"Col{i+1}" for i in range(col_count)]
        body = rows
    else:
        if header_row is None:
            header_idx = next((i for i, r in enumerate(rows) if any(str(c).strip() != "" for c in r)), 0)
        else:
            header_idx = max(0, min(len(rows) - 1, header_row - 1))
        header = rows[header_idx]
        body = rows[header_idx + 1:]

    header = [_escape_md_cell(x) for x in header]
    body = [[_escape_md_cell(x) for x in r] for r in body]
    sep = ["---"] * col_count

    lines = []
    lines.append("| " + " | ".join(header) + " |")
    lines.append("| " + " | ".join(sep) + " |")
    for r in body:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)


def _select_sheet_names(wb, sheets=None, sheet_indexes=None, sheet_regex=None):
    sheetnames = list(wb.sheetnames)
    targets = []

    if sheet_indexes:
        for idx in sheet_indexes:
            if 1 <= idx <= len(sheetnames):
                targets.append(sheetnames[idx - 1])
            else:
                raise ValueError(f"sheet_index_out_of_range: {idx}")

    if sheets:
        for name in sheets:
            if name in sheetnames:
                targets.append(name)
            else:
                raise ValueError(f"sheet_not_found: {name}")

    if sheet_regex:
        rx = re.compile(sheet_regex)
        for name in sheetnames:
            if rx.search(name):
                targets.append(name)

    if not targets:
        targets = sheetnames

    seen = set()
    out = []
    for n in targets:
        if n in seen:
            continue
        seen.add(n)
        out.append(n)
    return out


def _heading(level, title):
    lvl = max(1, int(level))
    return ("#" * lvl) + " " + title


def excel_to_markdown(
    file_path,
    sheets=None,
    sheet_indexes=None,
    sheet_regex=None,
    formulas=False,
    trim_outer=True,
    header_row=None,
    no_header=False,
    max_cells=20000,
    max_rows=None,
    max_cols=None,
    split_tables=False,
    blank_rows_gap=1,
    min_table_rows=2,
    heading_level=2,
    no_headings=False,
):
    path = Path(file_path)
    if path.suffix.lower() == ".xls":
        raise ValueError("unsupported_format: .xls (please convert to .xlsx)")
    wb = openpyxl.load_workbook(path, data_only=(not formulas))

    targets = _select_sheet_names(wb, sheets=sheets, sheet_indexes=sheet_indexes, sheet_regex=sheet_regex)
    multi_sheet = len(targets) > 1

    parts = []
    for name in targets:
        ws = wb[name]
        matrix = _sheet_to_matrix(
            ws,
            trim_outer=trim_outer,
            max_cells=max_cells,
            max_rows=max_rows,
            max_cols=max_cols,
        )
        if not matrix:
            continue

        tables = (
            _split_tables_by_blank_rows(matrix, blank_rows_gap=blank_rows_gap, min_rows=min_table_rows)
            if split_tables
            else [matrix]
        )
        sheet_parts = []
        for i, t in enumerate(tables, start=1):
            md_table = _matrix_to_markdown_table(t, header_row=header_row, no_header=no_header)
            if not md_table.strip():
                continue
            if no_headings:
                sheet_parts.append(md_table)
            else:
                if multi_sheet:
                    if len(tables) > 1:
                        sheet_parts.append(_heading(heading_level + 1, f"{name} - Table {i}") + "\n\n" + md_table)
                    else:
                        sheet_parts.append(_heading(heading_level, name) + "\n\n" + md_table)
                else:
                    if len(tables) > 1:
                        sheet_parts.append(_heading(heading_level, f"Table {i}") + "\n\n" + md_table)
                    else:
                        sheet_parts.append(md_table)
        if sheet_parts:
            if multi_sheet and not no_headings and len(tables) == 1:
                parts.append(sheet_parts[0])
            else:
                parts.extend(sheet_parts)

    return "\n\n".join([p for p in parts if p.strip() != ""])


def _build_parser():
    p = argparse.ArgumentParser(prog="excel_to_markdown.py")
    p.add_argument("input", help="Excel file path (.xlsx)")
    p.add_argument("-o", "--output", help="Output markdown file path (default: stdout)")
    p.add_argument("--sheet", action="append", help="Sheet name to export (repeatable)")
    p.add_argument("--sheet-index", action="append", type=int, help="1-based sheet index to export (repeatable)")
    p.add_argument("--sheet-regex", help="Regex to select sheets by name")
    p.add_argument("--formulas", action="store_true", help="Export formulas instead of calculated values")
    p.add_argument("--no-trim", action="store_true", help="Do not trim outer empty rows/cols")
    p.add_argument("--header-row", type=int, help="1-based header row index inside used range")
    p.add_argument("--no-header", action="store_true", help="Generate synthetic headers Col1..")
    p.add_argument("--max-cells", type=int, default=20000, help="Safety limit for sheet cells")
    p.add_argument("--max-rows", type=int, help="Limit exported rows (after trim)")
    p.add_argument("--max-cols", type=int, help="Limit exported cols (after trim)")
    p.add_argument("--split-tables", action="store_true", help="Split multiple tables by blank rows")
    p.add_argument("--blank-rows-gap", type=int, default=1, help="Blank rows gap to split tables")
    p.add_argument("--min-table-rows", type=int, default=2, help="Minimum rows to keep a table")
    p.add_argument("--heading-level", type=int, default=2, help="Base heading level when emitting headings")
    p.add_argument("--no-headings", action="store_true", help="Do not emit any headings")
    return p


def main(argv=None):
    argv = list(sys.argv[1:] if argv is None else argv)
    args = _build_parser().parse_args(argv)

    md = excel_to_markdown(
        args.input,
        sheets=args.sheet,
        sheet_indexes=args.sheet_index,
        sheet_regex=args.sheet_regex,
        formulas=args.formulas,
        trim_outer=(not args.no_trim),
        header_row=args.header_row,
        no_header=args.no_header,
        max_cells=args.max_cells,
        max_rows=args.max_rows,
        max_cols=args.max_cols,
        split_tables=args.split_tables,
        blank_rows_gap=args.blank_rows_gap,
        min_table_rows=args.min_table_rows,
        heading_level=args.heading_level,
        no_headings=args.no_headings,
    )

    if args.output:
        Path(args.output).write_text(md + ("" if md.endswith("\n") else "\n"), encoding="utf-8")
    else:
        sys.stdout.write(md)
        if md and not md.endswith("\n"):
            sys.stdout.write("\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
